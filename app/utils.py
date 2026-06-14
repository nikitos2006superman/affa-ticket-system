"""Вспомогательные функции: генерация QR и экспорт отчётов в Excel."""

import io
from datetime import datetime

import qrcode
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def make_qr_png(payload: str) -> bytes:
    """Возвращает PNG-байты QR-кода для строки payload."""
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=8,
        border=2,
    )
    qr.add_data(payload)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf.read()


# =====================================================================
# Excel-экспорт отчётов
# =====================================================================

_HEADER_FONT = Font(bold=True, color='FFFFFF')
_HEADER_FILL = PatternFill('solid', fgColor='2E75B6')


def _apply_header_style(ws, columns):
    for col_idx, _ in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _autosize(ws, columns):
    for col_idx, _ in enumerate(columns, start=1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)


def export_event_sales_xlsx(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Продажи по событиям'

    columns = [
        ('Мероприятие',     'title'),
        ('Дата',            'starts_at'),
        ('Статус',          'status'),
        ('Всего мест',      'total_capacity'),
        ('Продано',         'sold_count'),
        ('Свободно',        'available'),
        ('Заполненность %', 'fill_rate_pct'),
        ('Активных',        'tickets_active'),
        ('Использовано',    'tickets_used'),
        ('Возврат',         'tickets_cancelled'),
        ('Выручка',         'revenue'),
    ]
    ws.append([c[0] for c in columns])
    _apply_header_style(ws, columns)

    for r in rows:
        row = []
        for _, key in columns:
            val = r[key]
            if isinstance(val, datetime):
                val = val.strftime('%Y-%m-%d %H:%M')
            row.append(val)
        ws.append(row)

    _autosize(ws, columns)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_daily_revenue_xlsx(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Выручка по дням'

    columns = [
        ('Дата',                'sale_date'),
        ('Билетов продано',     'tickets_sold'),
        ('Уникальных покупателей', 'unique_buyers'),
        ('Событий с продажами', 'events_with_sales'),
        ('Валовая выручка',     'gross_revenue'),
        ('Сумма возвратов',     'refunded_amount'),
        ('Чистая выручка',      'net_revenue'),
    ]
    ws.append([c[0] for c in columns])
    _apply_header_style(ws, columns)

    for r in rows:
        row = []
        for _, key in columns:
            val = r[key]
            if hasattr(val, 'strftime'):
                val = val.strftime('%Y-%m-%d')
            row.append(val)
        ws.append(row)

    _autosize(ws, columns)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_attendance_xlsx(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Посещаемость'

    columns = [
        ('Мероприятие',           'title'),
        ('Дата',                  'starts_at'),
        ('Билетов выпущено',      'tickets_issued'),
        ('Пришло',                'attended'),
        ('Не пришло',             'no_shows'),
        ('Посещаемость, %',       'attendance_rate_pct'),
    ]
    ws.append([c[0] for c in columns])
    _apply_header_style(ws, columns)

    for r in rows:
        row = []
        for _, key in columns:
            val = r[key]
            if isinstance(val, datetime):
                val = val.strftime('%Y-%m-%d %H:%M')
            row.append(val)
        ws.append(row)

    _autosize(ws, columns)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_dealer_stats_xlsx(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Активность дилеров'

    columns = [
        ('Дилерский код',     'dealer_code'),
        ('Организация',       'organization'),
        ('Сотрудников',       'users_count'),
        ('Регистраций',       'tickets_total'),
        ('Посещений',         'tickets_attended'),
        ('Возвратов',         'tickets_cancelled'),
        ('Уник. событий',     'events_visited'),
        ('Посещаемость, %',   'attendance_rate_pct'),
    ]
    ws.append([c[0] for c in columns])
    _apply_header_style(ws, columns)

    for r in rows:
        row = []
        for _, key in columns:
            val = r[key]
            row.append(val)
        ws.append(row)

    _autosize(ws, columns)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
